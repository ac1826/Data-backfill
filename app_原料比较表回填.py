import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

st.set_page_config(page_title='各厂原料比较表回填', layout='wide')
st.title('各厂原料比较表回填')

DEFAULT_SOURCE_DIR = r'F:\llqdocument\大成文件\熟食专案'
DEFAULT_TARGET_FILE = r'F:\llqdocument\大成文件\熟食专案\系统成本_1\各厂原料比较表2603VSQ4 .xlsx'
DEFAULT_MAPPING_TEXT = '\n'.join([
    'BB=BB09',
    'BB2=BB49',
    'TJ=TJ09',
    'LY=LY05',
    'DL=GC11',
])
DEFAULT_SOURCE_SHEET_NAME = ''
DEFAULT_SOURCE_COLS = 'A-F,G-J'
DEFAULT_TARGET_COLS = 'A-F,H-K'
DEFAULT_FILTER_COL = 'A'
DEFAULT_EXCLUDE_PREFIXES = '3900,P,J'
DEFAULT_START_ROW = 2
DEFAULT_CLEAR_OLD_ROWS = True


def extract_month_code(file_name: str) -> str:
    matches = re.findall(r'(\d{4})', str(file_name))
    return matches[0] if matches else ''


def parse_mapping_text(text: str):
    mappings = []
    for raw_line in str(text).splitlines():
        line = raw_line.strip()
        if not line or line.startswith('#'):
            continue
        if '->' in line:
            left, right = line.split('->', 1)
        elif '=' in line:
            left, right = line.split('=', 1)
        else:
            raise ValueError(f'映射格式错误：{line}，请使用 BB=BB09 或 BB->BB09')
        source_key = left.strip()
        target_sheet = right.strip()
        if not source_key or not target_sheet:
            raise ValueError(f'映射格式错误：{line}')
        mappings.append((source_key, target_sheet))
    if not mappings:
        raise ValueError('至少需要一组文件与 Sheet 的映射关系')
    return mappings


def parse_prefixes(text: str):
    if not str(text).strip():
        return tuple()
    items = []
    for part in re.split(r'[,，\n]+', str(text)):
        item = part.strip()
        if item:
            items.append(item)
    return tuple(items)


def parse_column_spec(spec: str):
    columns = []
    text = str(spec).upper().replace(' ', '')
    if not text:
        raise ValueError('列范围不能为空')
    for part in text.split(','):
        if not part:
            continue
        if '-' in part:
            start_col, end_col = part.split('-', 1)
            start_idx = column_index_from_string(start_col)
            end_idx = column_index_from_string(end_col)
            if start_idx > end_idx:
                raise ValueError(f'列范围错误：{part}')
            columns.extend(range(start_idx, end_idx + 1))
        else:
            columns.append(column_index_from_string(part))
    if not columns:
        raise ValueError(f'无法解析列范围：{spec}')
    return columns


def normalize_cell_value(value):
    return value


def source_key_from_filename(file_name: str, month_code: str):
    stem = Path(file_name).stem.strip()
    if not stem:
        return ''
    if month_code:
        suffix = f'_{month_code}'
        if stem.endswith(suffix):
            return stem[:-len(suffix)]
    if '_' in stem:
        return stem.split('_', 1)[0].strip()
    return stem


def build_uploaded_source_map(source_dir: Path, month_code: str):
    source_map = {}
    duplicates = []
    for path in sorted(source_dir.glob('*.xlsx')):
        if path.name.startswith('~$'):
            continue
        source_key = source_key_from_filename(path.name, month_code)
        if not source_key:
            continue
        if source_key in source_map:
            duplicates.append(source_key)
            continue
        source_map[source_key] = path
    if duplicates:
        dup_text = '、'.join(sorted(set(duplicates)))
        raise ValueError(f'上传的源文件存在重复前缀：{dup_text}')
    return source_map


def get_source_worksheet(workbook, source_sheet_name: str):
    if source_sheet_name:
        if source_sheet_name not in workbook.sheetnames:
            raise ValueError(f'源文件缺少工作表：{source_sheet_name}')
        return workbook[source_sheet_name]
    return workbook[workbook.sheetnames[0]]


def read_filtered_rows(source_file: Path, source_columns, exclude_prefixes, filter_column_index: int, source_sheet_name: str):
    wb = load_workbook(source_file, read_only=True, data_only=True)
    try:
        ws = get_source_worksheet(wb, source_sheet_name)
        rows = []
        kept_count = 0
        skipped_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            filter_value = ''
            if filter_column_index - 1 < len(row) and row[filter_column_index - 1] is not None:
                filter_value = str(row[filter_column_index - 1]).strip()
            if not filter_value:
                skipped_count += 1
                continue
            if exclude_prefixes and filter_value.startswith(exclude_prefixes):
                skipped_count += 1
                continue
            rows.append([
                normalize_cell_value(row[col_idx - 1]) if col_idx - 1 < len(row) else None
                for col_idx in source_columns
            ])
            kept_count += 1
        return rows, kept_count, skipped_count
    finally:
        wb.close()


def clear_target_range(ws, start_row: int, end_row: int, target_columns):
    if end_row < start_row:
        return
    for row_idx in range(start_row, end_row + 1):
        for col_idx in target_columns:
            ws.cell(row=row_idx, column=col_idx).value = None


def copy_row_styles(ws, source_row_idx: int, target_row_idx: int, target_columns):
    for col_idx in target_columns:
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)
        if source_cell.has_style:
            target_cell._style = source_cell._style
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()
        if source_cell.protection:
            target_cell.protection = source_cell.protection.copy()


def ensure_target_rows_style(ws, start_row: int, rows_needed: int, target_columns):
    if rows_needed <= 0:
        return
    if ws.max_row >= start_row:
        template_row = start_row
    else:
        template_row = 1
    last_needed_row = start_row + rows_needed - 1
    for row_idx in range(max(ws.max_row + 1, start_row), last_needed_row + 1):
        copy_row_styles(ws, template_row, row_idx, target_columns)


def write_rows_to_sheet(ws, rows, start_row: int, target_columns, clear_old_rows: bool):
    ensure_target_rows_style(ws, start_row, len(rows), target_columns)
    current_max_row = max(ws.max_row, start_row + len(rows) - 1)
    if clear_old_rows:
        clear_target_range(ws, start_row, current_max_row, target_columns)

    for offset, row in enumerate(rows):
        row_idx = start_row + offset
        for value, col_idx in zip(row, target_columns):
            ws.cell(row=row_idx, column=col_idx).value = value

    return len(rows)


def build_output_path(target_file: Path, month_code: str):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_month = month_code or extract_month_code(target_file.name) or 'unknown'
    return target_file.with_name(f'filled_output_{safe_month}_{ts}{target_file.suffix}')


def get_target_sheetnames(target_file: Path):
    wb = load_workbook(target_file, read_only=True, data_only=False)
    try:
        return set(wb.sheetnames)
    finally:
        wb.close()


def grouped_target_columns(target_columns):
    groups = []
    if not target_columns:
        return groups
    start_pos = 0
    start_col = target_columns[0]
    prev_col = start_col
    for pos, col_idx in enumerate(target_columns[1:], start=1):
        if col_idx == prev_col + 1:
            prev_col = col_idx
            continue
        groups.append((start_col, prev_col, start_pos, pos - 1))
        start_pos = pos
        start_col = col_idx
        prev_col = col_idx
    groups.append((start_col, prev_col, start_pos, len(target_columns) - 1))
    return groups


def excel_used_last_row(ws):
    used_range = ws.UsedRange
    return used_range.Row + used_range.Rows.Count - 1


def clear_excel_columns(ws, start_row: int, end_row: int, target_columns):
    if end_row < start_row:
        return
    for first_col, last_col, _start_pos, _end_pos in grouped_target_columns(target_columns):
        ws.Range(ws.Cells(start_row, first_col), ws.Cells(end_row, last_col)).ClearContents()


def write_rows_to_sheet_excel(ws, rows, start_row: int, target_columns, clear_old_rows: bool):
    rows_needed = len(rows)
    last_needed_row = start_row + rows_needed - 1
    current_max_row = max(excel_used_last_row(ws), last_needed_row)

    if clear_old_rows:
        clear_excel_columns(ws, start_row, current_max_row, target_columns)

    if rows_needed <= 0:
        return 0

    min_col = min(target_columns)
    max_col = max(target_columns)
    used_last_row = excel_used_last_row(ws)
    if last_needed_row > used_last_row and used_last_row >= start_row:
        # Preserve the visual template when source data has more rows than the workbook currently uses.
        ws.Range(ws.Cells(start_row, min_col), ws.Cells(start_row, max_col)).Copy()
        ws.Range(ws.Cells(used_last_row + 1, min_col), ws.Cells(last_needed_row, max_col)).PasteSpecial(Paste=-4122)

    for first_col, last_col, start_pos, end_pos in grouped_target_columns(target_columns):
        values = tuple(
            tuple(row[pos] if pos < len(row) else None for pos in range(start_pos, end_pos + 1))
            for row in rows
        )
        target_range = ws.Range(ws.Cells(start_row, first_col), ws.Cells(last_needed_row, last_col))
        target_range.Value = values

    return rows_needed


def excel_cell_has_formula(cell):
    try:
        return bool(cell.HasFormula)
    except Exception:
        return False


def get_summary_formula_last_row(ws, data_start_row=4):
    # Excel 365 dynamic array formulas expose SpillingToRange. Use it when possible.
    try:
        spill_range = ws.Range('A3').SpillingToRange
        if spill_range is not None:
            return max(data_start_row, spill_range.Row + spill_range.Rows.Count - 1)
    except Exception:
        pass

    used_last_row = excel_used_last_row(ws)
    for row_idx in range(used_last_row, data_start_row - 1, -1):
        value = ws.Cells(row_idx, 1).Value
        if value not in (None, ''):
            return row_idx
    return data_start_row


def repair_summary_formulas_excel(ws, data_start_row=4):
    last_row = get_summary_formula_last_row(ws, data_start_row=data_start_row)
    if last_row < data_start_row:
        return 0

    used_range = ws.UsedRange
    last_col = used_range.Column + used_range.Columns.Count - 1
    changed_cols = 0
    for col_idx in range(2, last_col + 1):
        top_cell = ws.Cells(data_start_row, col_idx)
        if excel_cell_has_formula(top_cell):
            ws.Range(ws.Cells(data_start_row, col_idx), ws.Cells(last_row, col_idx)).FillDown()
            changed_cols += 1
    return changed_cols


def fill_target_workbook_with_excel(target_file: Path, output_file: Path, rows_by_sheet, start_row: int, target_columns, clear_old_rows: bool):
    """Use real Excel instead of openpyxl so Excel 365 dynamic-array formulas are preserved."""
    try:
        import pythoncom
        import win32com.client as win32
    except Exception as exc:
        raise RuntimeError(
            '需要安装 pywin32 才能用 Excel 保留动态数组公式。请先运行：py -3.12 -m pip install pywin32'
        ) from exc

    shutil.copy2(target_file, output_file)

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Open(str(output_file), UpdateLinks=0, ReadOnly=False)
        written_by_sheet = {}
        repaired_summary_cols = {}

        for sheet_name, rows in rows_by_sheet.items():
            ws = workbook.Worksheets(sheet_name)
            written_by_sheet[sheet_name] = write_rows_to_sheet_excel(
                ws=ws,
                rows=rows,
                start_row=start_row,
                target_columns=target_columns,
                clear_old_rows=clear_old_rows,
            )

        try:
            excel.CalculateFullRebuild()
        except Exception:
            workbook.RefreshAll()
            excel.CalculateFull()

        for sheet_name in rows_by_sheet:
            summary_sheet_name = f'{sheet_name}汇总'
            try:
                ws_summary = workbook.Worksheets(summary_sheet_name)
            except Exception:
                continue
            repaired_summary_cols[summary_sheet_name] = repair_summary_formulas_excel(ws_summary)

        try:
            excel.CalculateFullRebuild()
        except Exception:
            excel.CalculateFull()

        workbook.Save()
        return written_by_sheet, repaired_summary_cols
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def is_valid_material_code(value):
    if value in (None, ''):
        return False
    text = str(value).strip()
    return text not in ('', '0')


def unique_material_codes_in_order(rows):
    seen = set()
    result = []
    for row in rows:
        if not row:
            continue
        code = row[0]
        if not is_valid_material_code(code):
            continue
        code_text = str(code).strip()
        if code_text in seen:
            continue
        seen.add(code_text)
        result.append(code_text)
    return result


def get_existing_summary_codes(ws, data_start_row=4):
    codes = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row_idx, column=1).value
        if is_valid_material_code(value):
            codes.append(str(value).strip())
    return codes


def find_last_summary_data_row(ws, data_start_row=4):
    last_row = data_start_row - 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row_idx, column=1).value
        if is_valid_material_code(value):
            last_row = row_idx
    return last_row


def find_summary_placeholder_row(ws, data_start_row=4):
    for row_idx in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row_idx, column=1).value
        if str(value).strip() == '0':
            return row_idx
    return None


def find_fill_formula(ws, source_row_idx: int, col_idx: int, max_lookback: int = 20):
    for row_idx in range(source_row_idx, max(0, source_row_idx - max_lookback), -1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(value, str) and value.startswith('='):
            return value, row_idx
    return None, None


def copy_full_row_with_translation(ws, source_row_idx: int, target_row_idx: int):
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)
        value = source_cell.value
        if isinstance(value, str) and value.startswith('='):
            target_cell.value = Translator(value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
        elif value in (None, ''):
            fallback_formula, formula_row_idx = find_fill_formula(ws, source_row_idx, col_idx)
            if fallback_formula:
                origin = ws.cell(row=formula_row_idx, column=col_idx).coordinate
                target_cell.value = Translator(fallback_formula, origin=origin).translate_formula(target_cell.coordinate)
            else:
                target_cell.value = value
        else:
            target_cell.value = value
        if source_cell.has_style:
            target_cell._style = source_cell._style
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = source_cell.font.copy()
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()
        if source_cell.border:
            target_cell.border = source_cell.border.copy()
        if source_cell.alignment:
            target_cell.alignment = source_cell.alignment.copy()
        if source_cell.protection:
            target_cell.protection = source_cell.protection.copy()
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height


def repair_summary_filldown(ws, data_start_row=4):
    placeholder_row = find_summary_placeholder_row(ws, data_start_row=data_start_row)
    if placeholder_row is None:
        last_data_row = find_last_summary_data_row(ws, data_start_row=data_start_row)
        if last_data_row < data_start_row:
            return 0
        placeholder_row = last_data_row + 1

    changed = 0
    for row_idx in range(data_start_row + 1, placeholder_row + 1):
        ws.row_dimensions[row_idx].height = ws.row_dimensions[row_idx - 1].height
        for col_idx in range(2, ws.max_column + 1):
            prev_cell = ws.cell(row=row_idx - 1, column=col_idx)
            curr_cell = ws.cell(row=row_idx, column=col_idx)
            prev_value = prev_cell.value
            if isinstance(prev_value, str) and prev_value.startswith('='):
                translated = Translator(prev_value, origin=prev_cell.coordinate).translate_formula(curr_cell.coordinate)
                if curr_cell.value != translated:
                    curr_cell.value = translated
                    changed += 1
                if prev_cell.has_style:
                    curr_cell._style = prev_cell._style
    return changed


def append_missing_materials_to_summary(wb, detail_sheet_name: str, detail_rows):
    summary_sheet_name = f'{detail_sheet_name}汇总'
    if summary_sheet_name not in wb.sheetnames:
        return 0, [], 0

    ws = wb[summary_sheet_name]
    repaired_cells = repair_summary_filldown(ws)
    detail_codes = unique_material_codes_in_order(detail_rows)
    existing_codes = set(get_existing_summary_codes(ws))
    missing_codes = [code for code in detail_codes if code not in existing_codes]

    if not missing_codes:
        return 0, [], repaired_cells

    last_data_row = find_last_summary_data_row(ws)
    if last_data_row < 4:
        raise ValueError(f'{summary_sheet_name} 未找到可下拉的模板数据行')

    placeholder_row = last_data_row + 1
    fill_source_row = last_data_row

    for index, material_code in enumerate(missing_codes):
        target_row = placeholder_row + index
        copy_full_row_with_translation(ws, fill_source_row, target_row)
        ws.cell(row=target_row, column=1).value = material_code
        fill_source_row = target_row

    new_placeholder_row = placeholder_row + len(missing_codes)
    copy_full_row_with_translation(ws, fill_source_row, new_placeholder_row)
    ws.cell(row=new_placeholder_row, column=1).value = 0

    return len(missing_codes), missing_codes, repaired_cells


with st.sidebar:
    st.header('上传文件')
    target_upload = st.file_uploader('上传目标文件', type=['xlsx'], accept_multiple_files=False)
    source_uploads = st.file_uploader('上传源文件（可多选）', type=['xlsx'], accept_multiple_files=True)
    run_btn = st.button('开始回填', type='primary')

if run_btn:
    errors = []
    logs = []
    month_code_text = ''
    source_sheet_name_text = DEFAULT_SOURCE_SHEET_NAME
    mapping_text = DEFAULT_MAPPING_TEXT
    source_cols_text = DEFAULT_SOURCE_COLS
    target_cols_text = DEFAULT_TARGET_COLS
    filter_col_text = DEFAULT_FILTER_COL
    exclude_prefix_text = DEFAULT_EXCLUDE_PREFIXES
    start_row = DEFAULT_START_ROW
    clear_old_rows = DEFAULT_CLEAR_OLD_ROWS

    try:
        mappings = parse_mapping_text(mapping_text)
    except Exception as exc:
        mappings = []
        errors.append(str(exc))

    try:
        source_columns = parse_column_spec(source_cols_text)
    except Exception as exc:
        source_columns = []
        errors.append(f'源表列范围错误：{exc}')

    try:
        target_columns = parse_column_spec(target_cols_text)
    except Exception as exc:
        target_columns = []
        errors.append(f'目标表列范围错误：{exc}')

    try:
        filter_column_index = column_index_from_string(filter_col_text.strip().upper())
    except Exception:
        filter_column_index = 1
        errors.append('过滤列格式错误，例如应填写 A')

    exclude_prefixes = parse_prefixes(exclude_prefix_text)

    if source_columns and target_columns and len(source_columns) != len(target_columns):
        errors.append('源表列数量与目标表列数量必须一致')

    if target_upload is None:
        errors.append('请先上传目标文件')
    if not source_uploads:
        errors.append('请先上传源文件')

    if errors:
        for msg in errors:
            st.error(msg)
    else:
        with tempfile.TemporaryDirectory() as tmp_dir_str:
            tmp_dir = Path(tmp_dir_str)
            source_dir = tmp_dir / 'sources'
            source_dir.mkdir(parents=True, exist_ok=True)
            target_dir = tmp_dir / 'target'
            target_dir.mkdir(parents=True, exist_ok=True)

            target_file = target_dir / target_upload.name
            target_file.write_bytes(target_upload.getbuffer())
            for uploaded in source_uploads:
                (source_dir / uploaded.name).write_bytes(uploaded.getbuffer())

            month_code = month_code_text.strip() or extract_month_code(target_file.name)
            if not month_code:
                st.error('无法从目标文件名提取月份编码，请手动填写，例如 2603')
                st.stop()

            try:
                uploaded_source_map = build_uploaded_source_map(source_dir, month_code)
            except Exception as exc:
                st.error(str(exc))
                st.stop()

            try:
                target_sheetnames = get_target_sheetnames(target_file)
                summary = []
                rows_by_sheet = {}

                for source_key, sheet_name in mappings:
                    source_file = uploaded_source_map.get(source_key)
                    if source_file is None:
                        logs.append(f'未找到源文件：{source_key}（月份 {month_code}），跳过 {sheet_name}')
                        summary.append({
                            '源标识': source_key,
                            '目标Sheet': sheet_name,
                            '源文件': '未找到',
                            '保留行数': 0,
                            '过滤行数': 0,
                            '汇总新增物料号': 0,
                            '状态': '跳过',
                        })
                        continue

                    if sheet_name not in target_sheetnames:
                        logs.append(f'目标文件缺少 Sheet：{sheet_name}')
                        summary.append({
                            '源标识': source_key,
                            '目标Sheet': sheet_name,
                            '源文件': str(source_file.name),
                            '保留行数': 0,
                            '过滤行数': 0,
                            '汇总新增物料号': 0,
                            '状态': '目标Sheet不存在',
                        })
                        continue

                    summary_sheet_name = f'{sheet_name}汇总'
                    if summary_sheet_name not in target_sheetnames:
                        logs.append(f'目标文件缺少 Sheet：{summary_sheet_name}，仅回填明细 {sheet_name}')

                    rows, kept_count, skipped_count = read_filtered_rows(
                        source_file=source_file,
                        source_columns=source_columns,
                        exclude_prefixes=exclude_prefixes,
                        filter_column_index=filter_column_index,
                        source_sheet_name=source_sheet_name_text.strip(),
                    )
                    rows_by_sheet[sheet_name] = rows

                    logs.append(
                        f'{source_key} -> {sheet_name}：'
                        f'读取 {source_file.name}，保留 {kept_count} 行，过滤 {skipped_count} 行'
                    )

                    summary.append({
                        '源标识': source_key,
                        '目标Sheet': sheet_name,
                        '源文件': str(source_file.name),
                        '保留行数': kept_count,
                        '过滤行数': skipped_count,
                        '汇总新增物料号': '由UNIQUE公式自动生成',
                        '状态': '待读写',
                    })

                output_file = build_output_path(target_file, month_code)
                written_by_sheet, repaired_summary_cols = fill_target_workbook_with_excel(
                    target_file=target_file,
                    output_file=output_file,
                    rows_by_sheet=rows_by_sheet,
                    start_row=int(start_row),
                    target_columns=target_columns,
                    clear_old_rows=clear_old_rows,
                )
                for item in summary:
                    sheet_name = item.get('目标Sheet')
                    if sheet_name in written_by_sheet:
                        item['写入行数'] = written_by_sheet[sheet_name]
                        item['状态'] = '完成'
                for sheet_name, written in written_by_sheet.items():
                    logs.append(f'{sheet_name}：已用 Excel 写入 {written} 行，保留原工作簿公式/动态数组')
                for summary_sheet_name, fixed_cols in repaired_summary_cols.items():
                    logs.append(f'{summary_sheet_name}：已用 Excel 重新计算 UNIQUE，并下拉/补齐 {fixed_cols} 列公式')
                output_bytes = output_file.read_bytes()

                st.success('处理完成，可以直接下载结果文件')
                st.download_button(
                    label='下载结果文件',
                    data=output_bytes,
                    file_name=output_file.name,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                st.subheader('执行日志')
                for line in logs:
                    st.write('- ' + line)
                st.subheader('处理结果')
                st.dataframe(summary, use_container_width=True)
            except Exception as exc:
                st.exception(exc)
